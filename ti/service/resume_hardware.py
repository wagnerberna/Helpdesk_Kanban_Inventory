import os
import re
from os import listdir

import openpyxl
import pandas
from decouple import config
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side


class ProcessHardwareFiles:
    def __init__(self):
        self.file_name = "doc/resume.xlsx"
        self.path_files = "p:/ti/scriptbat/resume/"
        self.name_computers_process = []
        self.name_computers_error = []
        self.data_computers = []
        self.sector = ""
        self.user_name = ""
        self.computer_name = ""
        self.computer_type = ""
        self.ranking = ""
        self.cpu_manufacturer = ""
        self.cpu_model = ""
        self.cpu_generation = ""
        self.cpu_description = ""
        self.memory = ""
        self.hd_size = ""
        self.os = ""
        self.os_architecture = ""
        self.manufacturer = ""
        self.model = ""
        self.serial_number = ""

    def clear_fields(self):
        self.sector = ""
        self.user_name = ""
        self.computer_name = ""
        self.computer_type = ""
        self.ranking = ""
        self.cpu_manufacturer = ""
        self.cpu_model = ""
        self.cpu_generation = ""
        self.cpu_description = ""
        self.memory = ""
        self.hd_size = ""
        self.os = ""
        self.os_architecture = ""
        self.manufacturer = ""
        self.model = ""
        self.serial_number = ""

    def create_file_xls(self):
        if os.path.exists(self.file_name):
            os.remove(self.file_name)

        workbook = openpyxl.Workbook()
        workbook.save(self.file_name)

    def write_file(self, data):
        workbook = openpyxl.load_workbook(self.file_name)
        worksheet = workbook["Sheet"]
        row = len(worksheet["A"]) + 1

        column_a = "A" + str(row)
        column_b = "B" + str(row)
        column_c = "C" + str(row)
        column_d = "D" + str(row)
        column_e = "E" + str(row)
        column_f = "F" + str(row)
        column_g = "G" + str(row)
        column_h = "H" + str(row)
        column_i = "I" + str(row)
        column_j = "J" + str(row)
        column_k = "K" + str(row)
        column_l = "L" + str(row)
        column_m = "M" + str(row)
        column_n = "N" + str(row)
        column_o = "O" + str(row)
        column_p = "P" + str(row)

        worksheet[column_b] = data[0]
        worksheet[column_a] = data[1]
        worksheet[column_c] = data[2]
        worksheet[column_d] = data[3]
        worksheet[column_e] = data[4]
        worksheet[column_f] = data[5]
        worksheet[column_g] = data[6]
        worksheet[column_h] = data[7]
        worksheet[column_i] = data[8]
        worksheet[column_j] = data[9]
        worksheet[column_k] = data[10]
        worksheet[column_l] = data[11]
        worksheet[column_m] = data[12]
        worksheet[column_n] = data[13]
        worksheet[column_o] = data[14]
        worksheet[column_p] = data[15]

        workbook.save(self.file_name)

    def create_header(self):
        title_rows = [
            "Setor",
            "Usuario",
            "Computador",
            "Tipo",
            "Ranking",
            "CPU_Fabricante",
            "CPU_Modelo",
            "CPU_Geracao",
            "CPU_Descricao",
            "Memoria",
            "HD",
            "SO",
            "SO_bits",
            "Fabricante",
            "Modelo",
            "Serial",
        ]
        self.write_file(title_rows)

    def format_header(self):
        workbook = openpyxl.load_workbook(self.file_name)
        worksheet = workbook["Sheet"]
        worksheet.delete_rows(1)
        worksheet.column_dimensions["A"].width = 20
        worksheet.column_dimensions["B"].width = 12
        worksheet.column_dimensions["C"].width = 12
        worksheet.column_dimensions["D"].width = 12
        worksheet.column_dimensions["E"].width = 12
        worksheet.column_dimensions["F"].width = 12
        worksheet.column_dimensions["G"].width = 12
        worksheet.column_dimensions["H"].width = 12
        worksheet.column_dimensions["I"].width = 20
        worksheet.column_dimensions["J"].width = 12
        worksheet.column_dimensions["K"].width = 12
        worksheet.column_dimensions["L"].width = 20
        worksheet.column_dimensions["M"].width = 12
        worksheet.column_dimensions["N"].width = 12
        worksheet.column_dimensions["O"].width = 25
        worksheet.column_dimensions["P"].width = 12

        font_bold = Font(bold=True)
        header = [
            "A1",
            "B1",
            "C1",
            "D1",
            "E1",
            "F1",
            "G1",
            "H1",
            "I1",
            "J1",
            "K1",
            "L1",
            "M1",
            "N1",
            "O1",
            "P1",
        ]
        # worksheet["B5"].font = font_bold
        for cell in header:
            worksheet[cell].font = font_bold
        workbook.save(self.file_name)

    def format_border(self):
        workbook = openpyxl.load_workbook(self.file_name)
        worksheet = workbook["Sheet"]

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for row in worksheet:
            for cell in row:
                cell.border = thin_border

        workbook.save(self.file_name)

    def clean_user_name(self, data):
        domain = config("domain")
        find_user_name = list(filter(lambda el: "UserName=" in el, data))
        if not find_user_name:
            self.user_name = ""
        else:
            user_name = find_user_name[0]
            user_name = user_name.strip("UserName=" + domain).strip("\\").strip("\n")
            self.user_name = user_name.strip("\n")

    def check_user_sector(self, user):
        user_format = user.lower()
        data = pandas.read_excel("doc/Lista_usuarios_AD.xlsx")
        for index, row in data.iterrows():
            # print(user, index, row["login"], row["OU_1"])
            if row["login"] == user_format:
                self.sector = row["OU_1"]

        if user == "apontamento" or user == "metrologia":
            self.sector = "Fabrica"

    def clean_computer_name(self, data):
        find_computer_name = list(filter(lambda el: "Name=" in el, data))
        if not find_computer_name:
            self.computer_name = ""
        else:
            computer_name = find_computer_name[1]
            self.computer_name = computer_name.strip("Name=").strip("\n")

    def check_computer_type(self):
        if "N" in self.computer_name:
            self.computer_type = "Notebook"
        if "D" in self.computer_name:
            self.computer_type = "Desktop"

    def clean_cpu_description(self, data):
        find_computer_cpu = list(filter(lambda el: "Name=" in el, data))
        if not find_computer_cpu:
            self.cpu_description = ""
        else:
            cpu_description = find_computer_cpu[2]
            self.cpu_description = cpu_description.strip("Name=").strip("\n")

    def clean_hd_size(self, data):
        find_hd_size = list(filter(lambda el: "Size=" in el, data))
        if not find_hd_size:
            self.hd_size = ""
        else:
            hd_size = find_hd_size[0]
            hd_size = int(hd_size.strip("Size=").strip("\n"))
            converter = 1024**3
            self.hd_size = round(hd_size / converter)

    def clean_memory(self, data):
        find_memory = list(filter(lambda el: "TotalPhysicalMemory=" in el, data))
        if not find_memory:
            self.memory = ""
        else:
            memory = find_memory[0]
            memory = int(memory.strip("TotalPhysicalMemory=").strip("\n"))
            converter = 1024**2
            self.memory = round(memory / converter)

    def clean_os(self, data):
        find_os = list(filter(lambda el: "Caption=" in el, data))
        if not find_os:
            self.os = ""
        else:
            os_details = find_os[0]
            self.os = os_details.strip("Caption=Microsoft ").strip("\n")

    def clean_os_architecture(self, data):
        find_os_architecture = list(filter(lambda el: "OSArchitecture=" in el, data))
        if not find_os_architecture:
            self.os_architecture = ""
        else:
            os_architecture = find_os_architecture[0]
            self.os_architecture = os_architecture.strip("OSArchitecture=").strip("\n")

    def clean_manufacturer(self, data):
        find_manufacturer = list(filter(lambda el: "Manufacturer=" in el, data))
        if not find_manufacturer:
            self.manufacturer = ""
        else:
            manufacturer = find_manufacturer[0]
            manufacturer = manufacturer.strip("Manufacturer=").strip("\n").strip()
            if "Dell" in manufacturer:
                self.manufacturer = "Dell"
            elif "Hewlett" in manufacturer:
                self.manufacturer = "HP"
            elif "Hewlett" in manufacturer:
                self.manufacturer = "HP"
            elif "VAIO" in manufacturer:
                self.manufacturer = "VAIO"
            elif "Positivo" in manufacturer:
                self.manufacturer = "Positivo"
            else:
                self.manufacturer = manufacturer

    def clean_model(self, data):
        find_model = list(filter(lambda el: "Model=" in el, data))
        if not find_model:
            self.model = ""
        else:
            model = find_model[0]
            self.model = model.strip("Model=").strip("\n").strip()

    def clean_serial(self, data):
        find_serial = list(filter(lambda el: "SerialNumber=" in el, data))
        if not find_serial:
            self.serial_number = ""
        else:
            serial = find_serial[0]
            self.serial_number = serial.strip("SerialNumber=").strip("\n")

    def cpu_all_fields(self):
        if not self.cpu_description:
            self.cpu_manufacturer = ""

        else:
            if "Intel" in self.cpu_description:
                self.cpu_manufacturer = "Intel"
            if "Celeron" in self.cpu_description:
                self.cpu_model = "Celeron"
                self.ranking = "E"
            elif "Duo" in self.cpu_description:
                self.cpu_model = "Duo"
                self.ranking = "D"
            elif "Quad" in self.cpu_description:
                self.cpu_model = "Quad"
                self.ranking = "D"
            elif "Xeon" in self.cpu_description:
                self.cpu_model = "Xeon"
                self.ranking = "C"
            elif "i3" in self.cpu_description:
                self.cpu_model = "i3"
                find_cpu_generation = re.search("i3(.+?)CPU", self.cpu_description)
                if not find_cpu_generation:
                    self.ranking = "C"
                else:
                    cpu_generation = find_cpu_generation.group(1)
                    cpu_generation = "".join(re.findall("\d+", cpu_generation))
                    self.cpu_generation = int(cpu_generation)
                    if self.cpu_generation >= 5000:
                        self.ranking = "C"
                    else:
                        self.ranking = "D"

            elif "i5" in self.cpu_description:
                self.cpu_model = "i5"
                find_cpu_generation = re.search("i5(.+?)CPU", self.cpu_description)
                if not find_cpu_generation:
                    find_cpu_generation = re.search("i5(.+?)@", self.cpu_description)

                cpu_generation = find_cpu_generation.group(1)
                cpu_generation = "".join(re.findall("\d+", cpu_generation))
                self.cpu_generation = int(cpu_generation)
                if self.cpu_generation >= 8000:
                    self.ranking = "B"
                else:
                    self.ranking = "C"

            elif "i7" in self.cpu_description:
                self.cpu_model = "i7"
                find_cpu_generation = re.search("i7(.+?)CPU", self.cpu_description)
                if not find_cpu_generation:
                    find_cpu_generation = re.search("i7(.+?)@", self.cpu_description)

                cpu_generation = find_cpu_generation.group(1)
                cpu_generation = "".join(re.findall("\d+", cpu_generation))
                self.cpu_generation = int(cpu_generation)
                self.cpu_model = "i7"
                self.ranking = "A"

        if "AMD" in self.cpu_description:
            self.cpu_manufacturer = "AMD"
            if "X4" in self.cpu_description:
                self.cpu_model = "X4"
                self.ranking = "D"
            elif "Quad" in self.cpu_description:
                self.cpu_model = "Quad"
                self.ranking = "D"
            elif "Core" in self.cpu_description:
                self.cpu_model = "Core"
                self.ranking = "D"

    def process_file(self):
        try:
            # message_map = "Favor mapear a pasta 'Pacotes' do servidor 'Pateta' como unidade de rede com a letra: 'P'."
            # message_press_enter = "Pressione enter para continuar..."
            # message_process_file = "Processando arquivos:"
            # message_finish = "Arquivos processados com sucesso"
            # print(message_map)
            # input(message_press_enter)
            # print(message_process_file)

            self.create_file_xls()
            self.create_header()

            files_list = listdir(self.path_files)

            for file in files_list:
                file_path = self.path_files + file
                # print(file)

                with open(file_path, "r", encoding="utf-16") as file:
                    data = file.readlines()
                    if len(data) != 0:
                        self.clear_fields()
                        self.clean_user_name(data)
                        self.check_user_sector(self.user_name)
                        self.clean_computer_name(data)
                        self.check_computer_type()
                        self.clean_cpu_description(data)
                        self.cpu_all_fields()
                        self.clean_hd_size(data)
                        self.clean_memory(data)
                        self.clean_os(data)
                        self.clean_os_architecture(data)
                        self.clean_manufacturer(data)
                        self.clean_model(data)
                        self.clean_serial(data)
                        file.close
                        self.name_computers_process.append(self.computer_name)

                    else:
                        self.name_computers_error.append(self.computer_name)

                    data_computer = [
                        self.sector,
                        self.user_name,
                        self.computer_name,
                        self.computer_type,
                        self.ranking,
                        self.cpu_manufacturer,
                        self.cpu_model,
                        self.cpu_generation,
                        self.cpu_description,
                        self.memory,
                        self.hd_size,
                        self.os,
                        self.os_architecture,
                        self.manufacturer,
                        self.model,
                        self.serial_number,
                    ]

                    self.write_file(data_computer)

            self.format_header()
            self.format_border()
            # os.startfile(self.file_name)

            # print("Lista de arquivos processados:", self.name_computers_process)
            # print("Total de arquivos processados:", len(self.name_computers_process))
            # print("Erro ao processar os seguintes arquivos:", self.name_computers_error)
            # input(message_finish)
        except Exception as error:
            print("Internal error:", error)
            raise


# if __name__ == "__main__":
#     convert_files = ProcessHardwareFiles()
#     convert_files.process_file()
